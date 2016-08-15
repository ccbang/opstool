#!/usr/bin/env python
#-*- coding:utf-8 -*-
#need environment:
#1)yum -y install gcc gcc-c++ libffi-devel python-devel openssl-devel  mysql-devel libxml2 libxml2-dev libxslt* zlib openssl ; and other [mysql mysql-devel]
#2)you ensure "mysql_config" in environment path can find
#3)python version should be >= 2.7
#4)mkdir -p /data/export_excel
#4)pip install openpyxl
#5)pip install mysql

import MySQLdb
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import time
import getpass
# it's different between workbook and worksheet
pass_r = getpass.getpass('Please input mysql password:')
#connect to mysql
conn = MySQLdb.connect(host="172.18.7.10",user="root",passwd=pass_r,db="test",use_unicode=1,charset='utf8')
cur = conn.cursor()

# query sql
sql = "select * from gw_app;"
try:
    data = cur.execute(sql)
except Exception,e:
    print Exception,":",e

# use fetchall method to get a list, some tuples in it
try:
    data_list = cur.fetchall()
except Exception,e:
    print Exception,":",e

fields = cur.description
cur.close()
conn.close()

wb = Workbook()
ws = wb.worksheets[0]

#worksheet name(you can set)
ws.title = 'data'

wbk = ExcelWriter(workbook = wb)

rows = len(data_list)
cols = len(data_list[0])
# same as read data

#only output data
#for rx in range(rows):
#    for cx in range(cols):
#        ws.cell(row=rx+1, column=cx+1).value = data_list[rx][cx]

#output title and data
for ifs in range(cols):
    ws.cell(row=1, column=ifs+1).value = fields[ifs][0]
for rx in range(1,rows+1):
    for cx in range(cols):
        ws.cell(row=rx+1, column=cx+1).value = data_list[rx-1][cx]

wbk.save('/data/export_excel/'+str(time.strftime("%Y%m%d%H%M"))+'.xlsx')
